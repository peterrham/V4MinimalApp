#!/usr/bin/env python3
"""
Merge 'Testing Box Inventory Authority.xlsx' into inventory.json.

Reads:
  - /tmp/inventory.json (pulled from device)
  - ~/Downloads/Testing Box Inventory Authority.xlsx

Writes:
  - /tmp/inventory_merged.json (ready to push back to device)

For each xlsx BoxItem row:
  - Tries to match to existing inventory item by name (case-insensitive substring)
  - If match found: sets container field on existing item
  - If no match: creates new InventoryItem with name, container, comment->notes

The "container" is built from BoxHeaders: "Box {BoxName}: {BoxDescription}" or
just the location name for room-based containers.
"""

import json
import uuid
import sys
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("ERROR: pip3 install openpyxl")
    sys.exit(1)

XLSX_PATH = "/Users/peterham/Downloads/Testing Box Inventory Authority.xlsx"
INVENTORY_PATH = "/tmp/inventory.json"
OUTPUT_PATH = "/tmp/inventory_merged.json"

def load_inventory():
    with open(INVENTORY_PATH) as f:
        return json.load(f)

def load_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH)

    # Build box descriptions
    box_desc = {}
    box_status = {}
    ws_headers = wb["BoxHeaders"]
    for row in ws_headers.iter_rows(min_row=2, max_row=ws_headers.max_row, values_only=True):
        box_name = str(row[0]).strip() if row[0] else ""
        desc = str(row[1]).strip() if row[1] and not isinstance(row[1], float) else ""
        status = str(row[2]).strip() if row[2] else ""
        if box_name:
            box_desc[box_name] = desc
            box_status[box_name] = status

    # Load items
    items = []
    ws_items = wb["BoxItems"]
    for row in ws_items.iter_rows(min_row=2, max_row=ws_items.max_row, values_only=True):
        box_name = str(row[0]).strip() if row[0] else ""
        item_name = str(row[1]).strip() if row[1] else ""
        comment = str(row[2]).strip() if row[2] else ""
        quantity = row[3] if isinstance(row[3], (int, float)) else None
        status = str(row[4]).strip() if row[4] else ""
        date = row[5]
        poland_trip = row[7]
        link = str(row[8]).strip() if row[8] else ""

        if not item_name:
            continue

        # Build container label
        desc = box_desc.get(box_name, "")
        if box_name.isdigit():
            container = f"Box {box_name}"
            if desc:
                container += f": {desc}"
        else:
            container = box_name

        items.append({
            "name": item_name,
            "container": container,
            "comment": comment,
            "quantity": int(quantity) if quantity else None,
            "status": status,
            "link": link,
        })

    return items, box_desc, box_status

def normalize(name):
    return name.lower().strip().rstrip("s")  # basic normalization

def is_garbage(name):
    """Filter out speech recognition artifacts and non-items."""
    lower = name.lower().strip()
    if len(lower) < 2:
        return True
    garbage = ["hello", "this is my iphone", "hi there", "okay", "hey"]
    return lower in garbage

def find_match(item_name, inventory_items):
    """Try to find an existing inventory item matching by name."""
    norm = normalize(item_name)
    if len(norm) < 3:
        return None

    for i, inv_item in enumerate(inventory_items):
        inv_norm = normalize(inv_item["name"])
        # Exact match
        if inv_norm == norm:
            return i
        # Substring match (both directions)
        if len(norm) >= 5 and (norm in inv_norm or inv_norm in norm):
            return i
    return None

def make_inventory_item(xlsx_item):
    """Create a new InventoryItem dict from xlsx data."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    notes_parts = []
    if xlsx_item["comment"]:
        notes_parts.append(xlsx_item["comment"])
    if xlsx_item["quantity"] and xlsx_item["quantity"] > 1:
        notes_parts.append(f"Qty: {xlsx_item['quantity']}")
    if xlsx_item["link"]:
        notes_parts.append(f"Link: {xlsx_item['link']}")

    return {
        "id": str(uuid.uuid4()).upper(),
        "name": xlsx_item["name"],
        "category": "Other",
        "room": "",
        "container": xlsx_item["container"],
        "estimatedValue": None,
        "purchasePrice": None,
        "purchaseDate": None,
        "brand": None,
        "itemColor": None,
        "size": None,
        "notes": "; ".join(notes_parts),
        "photos": [],
        "voiceTranscripts": [],
        "createdAt": now,
        "updatedAt": now,
    }

def main():
    print("Loading inventory.json...")
    inventory = load_inventory()
    print(f"  {len(inventory)} existing items")

    print("Loading xlsx...")
    xlsx_items, box_desc, box_status = load_xlsx()
    print(f"  {len(xlsx_items)} items from spreadsheet")
    print(f"  {len(box_desc)} box definitions")

    matched = 0
    created = 0
    skipped = 0

    for xi in xlsx_items:
        if is_garbage(xi["name"]):
            skipped += 1
            continue

        match_idx = find_match(xi["name"], inventory)
        if match_idx is not None:
            # Set container on existing item
            if not inventory[match_idx].get("container"):
                inventory[match_idx]["container"] = xi["container"]
            # Append comment to notes if useful
            if xi["comment"] and xi["comment"] not in inventory[match_idx].get("notes", ""):
                existing_notes = inventory[match_idx].get("notes", "")
                if existing_notes:
                    inventory[match_idx]["notes"] = existing_notes + "; " + xi["comment"]
                else:
                    inventory[match_idx]["notes"] = xi["comment"]
            matched += 1
        else:
            # Create new item
            new_item = make_inventory_item(xi)
            inventory.append(new_item)
            created += 1

    print(f"\nResults:")
    print(f"  Matched to existing: {matched}")
    print(f"  New items created:   {created}")
    print(f"  Skipped (garbage):   {skipped}")
    print(f"  Total items now:     {len(inventory)}")

    # Container stats
    with_container = sum(1 for i in inventory if i.get("container"))
    print(f"  Items with container: {with_container}")

    # Write output
    with open(OUTPUT_PATH, "w") as f:
        json.dump(inventory, f, indent=2, default=str)
    print(f"\nWrote {OUTPUT_PATH}")

    # Show sample of containers
    from collections import Counter
    containers = Counter(i.get("container", "") for i in inventory if i.get("container"))
    print(f"\nTop 15 containers:")
    for c, count in containers.most_common(15):
        print(f"  {count:4}  {c[:60]}")

if __name__ == "__main__":
    main()
